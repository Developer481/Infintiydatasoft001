from selenium import webdriver as wd 
from selenium.webdriver.common.keys import Keys as ky
from selenium.webdriver.common.by import By
from time import sleep
import pandas as pd
import openpyxl
import os

path = os.getcwd()
crf = path+ "\google.csv"
xs = crf.replace("\\\\","\\")

c1 = path+ r"\googleurl.xlsx"
cs = c1.replace("\\\\","\\")

dataframe = openpyxl.load_workbook(cs)
dataframe1 = dataframe.active
ck = dataframe1.max_row
ck = ck+1
count = 1

driver = wd.Chrome('chromedriver.exe')

scroll_pause_time = 1
screen_height = driver.execute_script("return window.screen.height;")   # get the screen height of the web
i = 1
cz = 0


lis1 = []
lis2 = []
lis3 = []
lis4 = []


number = 1
SCROLL_PAUSE_TIME = 10

#url_1 = "https://www.google.co.in/maps/search/hospital+mohali/@30.7025669,76.7203382,15z/data=!3m1!4b1?hl=en"


for j in range(2,ck):
    url_1 = dataframe1.cell(j,1).value
    driver.get(url_1)
    #for i in range(0,5):
    while True:
        try:
            number = number+1
            #find data
            ur=driver.find_elements(by=By.XPATH,value="//a[@class ='hfpxzc']")
            nam = ur[cz].get_attribute('aria-label')
            print(nam)
            lis1.append(nam)
            #sleep(3)
            sleep(1)
            con=driver.find_element(by=By.XPATH,value="//a[@aria-label = '" + nam + "']")
            con.click()
            #sleep(5)
            sleep(2)
            #find extra inforamtion
            c1 = driver.find_element(by=By.XPATH,value="//button[@data-tooltip = 'Copy address']")
            adree = c1.get_attribute('aria-label')
            lis4.append(adree)
            c2 = driver.find_element(by=By.XPATH,value="//a[@data-tooltip = 'Open website']")
            web = c2.get_attribute('href')
            lis3.append(web)
            #end
            con1 = driver.find_element(by=By.XPATH,value="//button[@data-tooltip = 'Copy phone number']")
            con2 = con1.get_attribute('aria-label')
            lis2.append(con2)
            #sleep(3)
            sleep(1)
            con3 = driver.find_element(by=By.XPATH,value="//button[@aria-label = 'Close']")
            con3.click()
            sleep(1)

            
        except IndexError:
            break
            
        except:
            '''if lis3 != lis1  :
                lis3.append("")

            if lis2 != lis1:
                lis2.append("")

            if lis4 != lis1:
                lis4.append("")'''
                
            #lis2.append("")
            continue
        
            
        finally:

            if len(lis3) != len(lis1):
                lis3.append("")

            if len(lis2) != len(lis1):
                lis2.append("")

            if len(lis4) != len(lis1):
                lis4.append("")
                
            # Scroll down to bottom
            ele = driver.find_element(by=By.XPATH,value="//*[@id='QA0Szd']/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]")
            driver.execute_script('arguments[0].scrollBy(0, 5000);', ele)
            # Wait to load page
            sleep(SCROLL_PAUSE_TIME)
            ele = driver.find_element(by=By.XPATH,value="//*[@id='QA0Szd']/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]")
            new_height = driver.execute_script("return arguments[0].scrollHeight", ele)
            last_height = new_height
            cz = cz+1

    

print(len(lis1))
print(len(lis2))
print(len(lis3))
print(len(lis4))


df = pd.DataFrame({'Name':lis1,'Contact no':lis2,'Address':lis4,'web':lis3})
df.to_csv(xs,index=False)


driver.quit()

    
