from selenium import webdriver as wd 
from selenium.webdriver.common.keys import Keys as ky
from selenium.webdriver.common.by import By
import openpyxl
from time import sleep
import pandas as pd
import os
import csv
from datetime import datetime
import gspread
import re
import threading

path = os.getcwd()
c1 = path+ r"\Linkedin.xlsx"
cs = c1.replace("\\\\","\\")
dataframe = openpyxl.load_workbook(cs)

def task(driver,name):
    global dataframe
    global cs
    url_1 = "http://linkedin.com/"
    driver.get(url_1)
    dataframe1 = dataframe.get_sheet_by_name("credential")
    dataframe2 = dataframe.get_sheet_by_name("Msg")
    #w1 = dataframe.get_sheet_by_name("Linked Profile Links")
    w1 = dataframe.get_sheet_by_name(name)
    
    #find id pass
    ic = dataframe1.cell(2,1).value
    pas = dataframe1.cell(2,2).value

    #find msg
    note = dataframe2.cell(2,1).value

    #find last row
    row_count = w1.max_row
    row_count = row_count+1


    ur=driver.find_element(by=By.XPATH,value="//input[@class ='input__input']")
    ur.send_keys(ic)
    ps=driver.find_element(by=By.XPATH,value="//input[@id ='session_password']")
    ps.send_keys(pas)
    bt=driver.find_element(by=By.XPATH,value="//button[@class ='sign-in-form__submit-button']")
    bt.click()
    pv = datetime.now()
    r1 = pv.date()
    print(row_count)
    for i in range(2,row_count):
        try:
            v1= w1.cell(i,1).value
            v2 =w1.cell(i,2).value
            if v2 == None:
                driver.execute_script("window.open('');")
                sleep(5)
                driver.switch_to.window(driver.window_handles[1])
                #print(v1)
                driver.get(v1)
                sleep(5)
                sk = driver.find_elements(by=By.XPATH,value="//button[@class ='artdeco-button artdeco-button--2 artdeco-button--primary ember-view']")
                vss = sk[1].get_attribute('id')
                ct = driver.find_element(by=By.XPATH,value="//button[@id = '" + vss + "']")
                ct.click()
                sleep(3)
                ms1 = driver.find_element(by=By.XPATH,value="//button[@aria-label= 'Add a note']")
                ms1.click()
                msg2 = driver.find_element(by=By.XPATH,value="//textarea[@id= 'custom-message']")
                msg2.send_keys(note)
                #sleep(3)
                hid =driver.find_element(by=By.XPATH,value="//button[@aria-label= 'Send now']")
                hid.click()
                sleep(2)
                #work with update version
                now = datetime.now()
                dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
                w1.cell(i,2).value = "yes"
                w1.cell(i,3).value = dt_string
                w1.cell(i,4).value = note
            else:
                nows = datetime.now()
                dt_strin = nows.strftime("%d/%m/%Y %H:%M:%S")
                #w1.update_cell(i,7,dt_strin)
                w1.cell(i,7).value = dt_strin
                driver.execute_script("window.open('');")
                sleep(5)
                driver.switch_to.window(driver.window_handles[1])
                driver.get(v1)
                sleep(5)
                tm=driver.find_elements(by=By.XPATH,value="//button[@class ='artdeco-button artdeco-button--muted artdeco-button--2 artdeco-button--secondary ember-view']")
                nam = tm[2].get_attribute("aria-label")
                if nam != None:
                    nam1 = nam[0:19]
                else:
                    nam1 = 'ok'
                if nam1 == 'Withdraw invitation':
                    dt_string = w1.cell(i,3).value
                    datetime_obj = datetime.strptime(dt_string,"%d/%m/%Y %H:%M:%S")
                    r2 =datetime_obj.date()
                    lvd= r1-r2
                    lvd1 = lvd.days
                    if lvd1 >30:
                        ts=driver.find_elements(by=By.XPATH,value="//button[@class ='artdeco-button artdeco-button--muted artdeco-button--2 artdeco-button--secondary ember-view']")
                        vt =ts[2].get_attribute("id")
                        ct = driver.find_element(by=By.XPATH,value="//button[@id = '" + vt + "']")
                        ct.click()
                        sleep(3)
                        tp=driver.find_element(by=By.XPATH,value="//button[@class ='artdeco-modal__confirm-dialog-btn artdeco-button artdeco-button--2 artdeco-button--primary ember-view']")
                        tk = tp.get_attribute("id")
                        tcs = driver.find_element(by=By.XPATH,value="//button[@id = '" + tk + "']")
                        tcs.click()
                        sleep(2)
                        #w1.update_cell(i,5,"withdraw")
                        w1.cell(i,5).value = "withdraw"
                    else:
                        #w1.update_cell(i,5,"Pending")
                        w1.cell(i,5).value = "Pending"
                else:
                    #w1.update_cell(i,5,"Accepted")
                    w1.cell(i,5).value = "Accepted"
                    if w1.cell(i,6).value == None :
                        svt1=driver.find_element(by=By.XPATH,value="//div[@class ='entry-point']")
                        svt =svt1.find_element(by=By.XPATH,value="//a[@class ='message-anywhere-button pvs-profile-actions__action artdeco-button ']")
                        msglink =svt.get_attribute('href')
                        driver.execute_script("window.open('about:blank', 'thirdtab');")
                        driver.switch_to.window("thirdtab")
                        driver.get(msglink)
                        sleep(5)
                        pst =driver.find_element(by=By.XPATH,value="//div[@aria-label= 'Write a message…']")
                        pst.send_keys(note)
                        #w1.update_cell(i,6,note)
                        w1.cell(i,6).value = note
                        btm =driver.find_element(by=By.XPATH,value="//button[@class= 'msg-form__send-button artdeco-button artdeco-button--1']")
                        btm.click()
                        driver.close()
                        driver.switch_to.window(driver.window_handles[1])
        except:
            now = datetime.now()
            dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
            #w1.update_cell(i,2,"no")
            w1.cell(i,2).value = "no"
            #w1.update_cell(i,3,dt_string)
            w1.cell(i,3).value =  dt_string
            #w1.update_cell(i,4,"")
            w1.cell(i,4).value = ""
            continue
        finally:
            driver.close()
            driver.switch_to.window(driver.window_handles[0])

    driver.quit()
    dataframe.save(cs)



    
    
ls = dataframe.get_sheet_names()
k = len(ls)
lg = k - 2
arval = 2

for tms in range(lg):
    sar = ls[arval]
    arval = arval+1
    driver = wd.Chrome('chromedriver.exe')
    t = threading.Thread(target=task, args=(driver,sar))
    t.start()
    
    











